"""Экспорт расписания в Excel (.xlsx) и PDF.

Оба формата строят одну и ту же сетку «урок × день недели», в ячейке —
название предмета, преподаватель, кабинет, класс (если не выбран) и
метка «закр.» для закреплённых занятий.

Источник данных — словарь grid, возвращаемый scheduler.views.build_week_grid().
"""

from __future__ import annotations

from dataclasses import dataclass
from datetime import date
from io import BytesIO
from typing import Iterable


@dataclass(frozen=True)
class ExportContext:
    """Минимум контекста для подписи документа."""
    week_start: date
    scope_label: str
    show_class_in_cell: bool  # True, если фильтр по классу НЕ задан


# ============================ ОБЩИЕ ХЕЛПЕРЫ ============================

def _lesson_cell_text(lesson, *, show_class_in_cell: bool) -> str:
    """Унифицированное текстовое представление урока для ячейки."""
    if lesson is None:
        return ''
    parts = [str(lesson.subject.name)]
    parts.append(str(lesson.teacher))
    parts.append(f'каб. {lesson.classroom.name}')
    if show_class_in_cell:
        parts.append(f'Класс {lesson.class_obj.name}')
    if lesson.is_locked:
        parts.append('закр.')
    return '\n'.join(parts)


def _filename_base(week_start: date) -> str:
    return f'raspisanie_{week_start.isoformat()}'


# ============================ EXCEL ============================

def build_xlsx(grid: dict, ctx: ExportContext) -> tuple[bytes, str]:
    """Строит файл .xlsx с расписанием.

    Возвращает (bytes, filename).
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = 'Расписание'

    # --- цвета и стили
    accent_fill = PatternFill('solid', fgColor='0F766E')
    header_fill = PatternFill('solid', fgColor='FBF9F4')
    soft_fill = PatternFill('solid', fgColor='F3F0EA')
    locked_fill = PatternFill('solid', fgColor='FFF4E0')

    title_font = Font(name='Calibri', size=16, bold=True, color='1C160E')
    sub_font = Font(name='Calibri', size=11, color='4A4036')
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    cell_font = Font(name='Calibri', size=10, color='1C160E')
    muted_font = Font(name='Calibri', size=10, italic=True, color='786E62')
    lesson_number_font = Font(name='Calibri', size=12, bold=True, color='1C160E')

    thin = Side(style='thin', color='D8D2C6')
    cell_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_wrap = Alignment(horizontal='left', vertical='top', wrap_text=True)

    weekdays = grid.get('weekdays', [])
    rows = grid.get('rows', [])
    total_cols = 1 + max(1, len(weekdays))

    # --- заголовок документа
    ws.cell(row=1, column=1, value='Школьное расписание').font = title_font
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    ws.row_dimensions[1].height = 26

    week_end = ctx.week_start
    if weekdays:
        week_end = weekdays[-1]['date']
    subtitle = (
        f'Неделя {ctx.week_start.strftime("%d.%m.%Y")} – '
        f'{week_end.strftime("%d.%m.%Y")} · {ctx.scope_label}'
    )
    ws.cell(row=2, column=1, value=subtitle).font = sub_font
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=total_cols)
    ws.row_dimensions[2].height = 20

    # --- шапка таблицы
    header_row = 4
    weekday_short = {1: 'Пн', 2: 'Вт', 3: 'Ср', 4: 'Чт', 5: 'Пт', 6: 'Сб', 7: 'Вс'}

    cell = ws.cell(row=header_row, column=1, value='Урок')
    cell.font = header_font
    cell.fill = accent_fill
    cell.alignment = center
    cell.border = cell_border

    for col_index, day in enumerate(weekdays, start=2):
        iso_weekday = day.get('weekday') or day['date'].isoweekday()
        label = f'{weekday_short.get(iso_weekday, "")} · {day["date"].strftime("%d.%m")}'
        cell = ws.cell(row=header_row, column=col_index, value=label)
        cell.font = header_font
        cell.fill = accent_fill
        cell.alignment = center
        cell.border = cell_border

    ws.row_dimensions[header_row].height = 32

    # --- содержимое
    body_start = header_row + 1
    for row_index, row in enumerate(rows):
        excel_row = body_start + row_index

        # Колонка «Урок»
        meta_cell = ws.cell(row=excel_row, column=1, value=str(row['lesson_number']))
        meta_cell.font = lesson_number_font
        meta_cell.alignment = center
        meta_cell.fill = soft_fill
        meta_cell.border = cell_border

        time_label = row.get('time') or ''
        if time_label:
            # время добавим в ту же ячейку через перенос строки
            meta_cell.value = f"{row['lesson_number']}\n{time_label}"
            meta_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # Колонки дней
        for col_index, day in enumerate(weekdays, start=2):
            lesson = (row.get('cells') or {}).get(day['date'])
            value = _lesson_cell_text(lesson, show_class_in_cell=ctx.show_class_in_cell)
            data_cell = ws.cell(row=excel_row, column=col_index, value=value or '—')
            data_cell.alignment = left_wrap
            data_cell.border = cell_border
            if lesson is None:
                data_cell.font = muted_font
                data_cell.alignment = center
            else:
                data_cell.font = cell_font
                if getattr(lesson, 'is_locked', False):
                    data_cell.fill = locked_fill
                else:
                    data_cell.fill = header_fill

        ws.row_dimensions[excel_row].height = 78 if any(
            (row.get('cells') or {}).get(day['date']) for day in weekdays
        ) else 38

    # --- ширины колонок
    ws.column_dimensions[get_column_letter(1)].width = 14
    for col_index in range(2, total_cols + 1):
        ws.column_dimensions[get_column_letter(col_index)].width = 28

    ws.sheet_view.showGridLines = False
    ws.freeze_panes = ws.cell(row=body_start, column=2)

    stream = BytesIO()
    wb.save(stream)
    return stream.getvalue(), f'{_filename_base(ctx.week_start)}.xlsx'


# ============================ PDF ============================

def build_pdf(grid: dict, ctx: ExportContext) -> tuple[bytes, str]:
    """Строит файл PDF (A4 landscape) с расписанием. Возвращает (bytes, filename)."""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from reportlab.platypus import (
        Paragraph,
        SimpleDocTemplate,
        Spacer,
        Table,
        TableStyle,
    )

    # Шрифт с кириллицей. DejaVuSans распространяется почти везде, где есть Linux/Win,
    # и обычно идёт в комплекте с matplotlib/reportlab. Пытаемся подключить, иначе fallback.
    font_name = 'Helvetica'
    font_bold = 'Helvetica-Bold'
    import os
    candidates = [
        '/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf',
        '/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf',
        'C:/Windows/Fonts/arial.ttf',
        'C:/Windows/Fonts/arialbd.ttf',
    ]
    try:
        regular_path = next((p for p in candidates if 'Bold' not in p and 'bd' not in p and os.path.exists(p)), None)
        bold_path = next((p for p in candidates if ('Bold' in p or 'bd' in p) and os.path.exists(p)), None)
        if regular_path:
            pdfmetrics.registerFont(TTFont('SchedulerSans', regular_path))
            font_name = 'SchedulerSans'
        if bold_path:
            pdfmetrics.registerFont(TTFont('SchedulerSans-Bold', bold_path))
            font_bold = 'SchedulerSans-Bold'
    except Exception:
        pass

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'TitleStyle', parent=styles['Title'],
        fontName=font_bold, fontSize=18, leading=22, textColor=colors.HexColor('#1C160E'),
        spaceAfter=4,
    )
    sub_style = ParagraphStyle(
        'SubStyle', parent=styles['Normal'],
        fontName=font_name, fontSize=10, leading=14, textColor=colors.HexColor('#4A4036'),
        spaceAfter=10,
    )
    cell_style = ParagraphStyle(
        'CellStyle', parent=styles['Normal'],
        fontName=font_name, fontSize=8, leading=10, textColor=colors.HexColor('#1C160E'),
    )
    cell_subject_style = ParagraphStyle(
        'CellSubject', parent=cell_style,
        fontName=font_bold, fontSize=8.5, leading=10.5,
    )
    cell_muted_style = ParagraphStyle(
        'CellMuted', parent=cell_style,
        textColor=colors.HexColor('#786E62'),
    )
    empty_style = ParagraphStyle(
        'EmptyStyle', parent=cell_style,
        textColor=colors.HexColor('#A39A8D'), alignment=1,
    )
    header_cell_style = ParagraphStyle(
        'HeaderCell', parent=cell_style,
        fontName=font_bold, fontSize=9, leading=11, textColor=colors.HexColor('#FFFFFF'),
        alignment=1,
    )
    lesson_number_style = ParagraphStyle(
        'LessonNumber', parent=cell_style,
        fontName=font_bold, fontSize=11, leading=13, alignment=1,
    )
    lesson_time_style = ParagraphStyle(
        'LessonTime', parent=cell_style,
        fontName=font_name, fontSize=7, leading=9,
        textColor=colors.HexColor('#786E62'), alignment=1,
    )

    weekdays = grid.get('weekdays', [])
    rows = grid.get('rows', [])
    weekday_short = {1: 'Пн', 2: 'Вт', 3: 'Ср', 4: 'Чт', 5: 'Пт', 6: 'Сб', 7: 'Вс'}

    # --- собираем таблицу
    head = [Paragraph('Урок', header_cell_style)]
    for day in weekdays:
        iso_weekday = day.get('weekday') or day['date'].isoweekday()
        head.append(Paragraph(
            f"{weekday_short.get(iso_weekday, '')} · {day['date'].strftime('%d.%m')}",
            header_cell_style,
        ))

    body_rows = []
    for row in rows:
        lesson_meta = [
            Paragraph(str(row['lesson_number']), lesson_number_style),
            Paragraph(row.get('time') or '', lesson_time_style),
        ]
        row_cells = [lesson_meta]
        for day in weekdays:
            lesson = (row.get('cells') or {}).get(day['date'])
            if lesson is None:
                row_cells.append(Paragraph('—', empty_style))
            else:
                parts = [Paragraph(str(lesson.subject.name), cell_subject_style)]
                parts.append(Paragraph(str(lesson.teacher), cell_style))
                parts.append(Paragraph(f'каб. {lesson.classroom.name}', cell_muted_style))
                if ctx.show_class_in_cell:
                    parts.append(Paragraph(f'Класс {lesson.class_obj.name}', cell_muted_style))
                if getattr(lesson, 'is_locked', False):
                    parts.append(Paragraph(
                        '<font color="#B3791F"><b>● закр.</b></font>',
                        cell_style,
                    ))
                row_cells.append(parts)
        body_rows.append(row_cells)

    data = [head] + body_rows

    page_w, page_h = landscape(A4)
    margin = 12 * mm
    usable_w = page_w - 2 * margin
    first_col_w = 22 * mm
    n_day_cols = max(1, len(weekdays))
    day_col_w = (usable_w - first_col_w) / n_day_cols
    col_widths = [first_col_w] + [day_col_w] * n_day_cols

    table = Table(data, colWidths=col_widths, repeatRows=1)
    table_style = TableStyle([
        # шапка
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0F766E')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('VALIGN', (0, 0), (-1, 0), 'MIDDLE'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 7),
        ('TOPPADDING', (0, 0), (-1, 0), 7),

        # тело
        ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#D8D2C6')),
        ('VALIGN', (0, 1), (-1, -1), 'TOP'),
        ('LEFTPADDING', (0, 1), (-1, -1), 5),
        ('RIGHTPADDING', (0, 1), (-1, -1), 5),
        ('TOPPADDING', (0, 1), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 5),

        # колонка «Урок»
        ('BACKGROUND', (0, 1), (0, -1), colors.HexColor('#F3F0EA')),
        ('VALIGN', (0, 1), (0, -1), 'MIDDLE'),
        ('ALIGN', (0, 1), (0, -1), 'CENTER'),

        # зебра по строкам
        ('ROWBACKGROUNDS', (1, 1), (-1, -1), [colors.white, colors.HexColor('#FBF9F4')]),
    ])
    table.setStyle(table_style)

    week_end = ctx.week_start
    if weekdays:
        week_end = weekdays[-1]['date']

    story = [
        Paragraph('Школьное расписание', title_style),
        Paragraph(
            f'Неделя {ctx.week_start.strftime("%d.%m.%Y")} – '
            f'{week_end.strftime("%d.%m.%Y")} · {ctx.scope_label}',
            sub_style,
        ),
        table,
        Spacer(1, 6 * mm),
        Paragraph(
            f'<font color="#A39A8D">Документ сгенерирован автоматически. '
            f'Закреплённые занятия отмечены значком ●.</font>',
            cell_muted_style,
        ),
    ]

    stream = BytesIO()
    doc = SimpleDocTemplate(
        stream,
        pagesize=landscape(A4),
        leftMargin=margin, rightMargin=margin,
        topMargin=margin, bottomMargin=margin,
        title='Школьное расписание',
        author='Школьное расписание',
    )
    doc.build(story)
    return stream.getvalue(), f'{_filename_base(ctx.week_start)}.pdf'


# ============================ ДИСПЕТЧЕР ============================

SUPPORTED_FORMATS = ('xlsx', 'pdf')

CONTENT_TYPES = {
    'xlsx': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    'pdf': 'application/pdf',
}


def export_grid(grid: dict, ctx: ExportContext, fmt: str) -> tuple[bytes, str, str]:
    """Универсальный диспетчер: возвращает (data, filename, content_type)."""
    fmt = (fmt or '').lower().strip()
    if fmt not in SUPPORTED_FORMATS:
        raise ValueError(f'Неподдерживаемый формат: {fmt!r}. Допустимые: {SUPPORTED_FORMATS}.')
    if fmt == 'xlsx':
        data, filename = build_xlsx(grid, ctx)
    else:
        data, filename = build_pdf(grid, ctx)
    return data, filename, CONTENT_TYPES[fmt]
